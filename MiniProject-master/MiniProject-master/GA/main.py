import TimeTable as tt
import assets.Attributes as att
import assets.AccesJsonDatas3 as sd3
import assets.AccesJsonDatas5 as sd5
import GeneticAlgorithm as ga
import numpy as np
from openpyxl import Workbook

def save_config():
    sd3.addfaculty(att.faculties)
    sd3.addsubject
    sd5.addsubject
def addfac():
    n = input("Enter name : ")
    i = input("enter id : ")
    obj = att.Faculty(n,i)
    att.Faculty.updatefac(obj)
    
def addsub(sem):
    na = input("Enter name : ")
    cod = input("Enter code : ")
    print(att.faculties)
    f = input("Choose a faculty by id : ")
    fac = att.faculties[f]
    m = input("Enter max hour : ")
    obj = att.Subject(na, cod, fac, m)
    if sem == 3 :
        att.sub3.append(obj.code)
        att.subjects3[obj.code]=obj
    elif sem == 5 :
        att.sub5.append(obj.code)
        att.subjects5[obj.code]=obj

def delfac():
    i = input("Enter id of faculty to remove : ")
    att.faculties.pop(i)
def delsub():
    # cod = input("Enter code of subject to remove : ")
    # att.subjects.pop(cod)
    pass
def scd():
    scggg = ga.generate_timetable(3)
    matrix = np.array(scggg)
    wb = Workbook()
    sheet = wb.active
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            sheet.cell(row=i+1, column=j+1, value=matrix[i, j])
    wb.save('output.xlsx')

def scd2():
    scggg = ga.generate_timetable(3)
    matrix = np.array(scggg)
    wb = Workbook()
    sheet = wb.active
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            sheet.cell(row=i+1, column=j+1, value=matrix[i, j])
    wb.save('output2.xlsx')

def semester3():
    while True:
        ch = int(input("1-add Faculty\n2-add Subject\n3-delete faculty\n4-delete subject\n5-exit\n6-schedule\n: "))
        if ch == 1 : addfac()
        elif ch == 2 : addsub(3)
        elif ch == 3 : delfac()
        elif ch == 4 : delsub()
        elif ch == 6 : scd()
        else : break

def semester5():
    while True:
        ch = int(input("1-add Faculty\n2-add Subject\n3-delete faculty\n4-delete subject\n5-exit\n6-schedule\n: "))
        if ch == 1 : addfac()
        elif ch == 2 : addsub(5)
        elif ch == 3 : delfac()
        elif ch == 4 : delsub()
        elif ch == 6 : scd2()
        else : break

print("trying to reload")
print(sd5.getsubjects)
while True:
    sem=int(input("1-sem3\n2-sem5\n3-exit\n :"))
    if sem == 1 : semester3()
    elif sem == 2 : semester5()
    else : break
