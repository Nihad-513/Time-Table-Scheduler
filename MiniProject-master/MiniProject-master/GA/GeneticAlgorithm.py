import GA.TimeTable as schedule
import random as r
from GA.assets import Attributes as att
import numpy as np
from openpyxl import Workbook
import mysql.connector as my

def xlsgenersting(sched1,sched2):
    matrix = np.array(sched1)
    wb = Workbook()
    sheet = wb.active
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            sheet.cell(row=i+1, column=j+1, value=matrix[i, j])
    wb.save('generated s3 time table.xlsx')
    matrix = np.array(sched2)
    wb = Workbook()
    sheet = wb.active
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            sheet.cell(row=i+1, column=j+1, value=matrix[i, j])
    wb.save('generated s5 time table.xlsx')

#generates a timetable for a given semester
def comgenerate(subje, sched):
    for i in sched:
        temp = subje.copy()
        print(subje)
        print(temp)
        r.shuffle(temp)
        for j in range(min(len(i), len(temp))):
            i[j] = temp.pop(0)
    return sched

def generate_timetable(semester):
    # num = schedule.num_hours
    sched = [[],[],[],[],[]]
    for i in range(5):
        sched[i]=[None]*6
    if semester == 'S3' : 
        mycon=my.connect(host='localhost',user='root',
                     passwd='mysql',database='tts')
        cur=mycon.cursor()
        qr1="select scode from subject where sem='S3'"
        cur.execute(qr1)
        lis=cur.fetchall()
        subje=[]
        for i in lis:
            subje.append(i[0])
        # print("S3 sinfd")
        # print(subje)
        # subje= att.sub3.copy()
        sched = comgenerate(subje,sched)
        schedule.generated_sem3 = sched.copy()
    elif semester == 'S5' :
        mycon=my.connect(host='localhost',user='root',
                     passwd='mysql',database='tts')
        cur=mycon.cursor()
        qr1="select scode from subject where sem='S5'"
        cur.execute(qr1)
        lis=cur.fetchall()
        subje=[]
        for i in lis:
            subje.append(i[0])
        # print("sunje")
        # print(subje)
        #subje= att.sub5.copy()
        sched = comgenerate(subje,sched)
        schedule.generated_sem5 = sched.copy()
        # print("sched",sched)    
    return sched

def fitness(sched1,sched2):
    mycon=my.connect(host='localhost',user='root',
                     passwd='mysql',database='tts')
    cur=mycon.cursor()
    for i in range(len(sched1)):
        for j in  range(len(sched1[0])):
            subject1 : str = sched1[1][3]
            subject2 : str = sched2[1][3]
            print(subject1,subject2)
            qr1="SELECT faculty FROM subject WHERE scode = %s"
            print(1)
            cur.execute(qr1, (subject1,))
            print(2)
            lis=cur.fetchone()
            print(3)
            for i in lis:
                faculty1 = i
            print("=======")
            print(faculty1)
            qr2="SELECT faculty FROM subject WHERE scode = %s"
            print(4)
            cur.execute(qr2, (subject2,))
            print(6)
            lis2=cur.fetchone()
            for i in lis2:
                faculty2 = i
            print(8)
            if faculty1 == faculty2:
                print('faculty conflict')
                fit1 = generate_timetable(3)
                fit2 = generate_timetable(5)
                fitness(fit1, fit2)
            print('1st run')
    else:

        # return schedule.generated_sem3,schedule.generated_sem5
        return schedule.generated_sem3,schedule.generated_sem5
    


def print_generated(schedu):
    for i in schedu:
        print(i)
    

